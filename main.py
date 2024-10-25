import telebot
from telebot import types
import openpyxl
import re
import time
from dates import *



bot = telebot.TeleBot('7655400381:AAFIYnMA_7HKJJoi7ight0gwsthjVlj630s')

workbook = openpyxl.load_workbook("24-knt.xlsx")
sheet = workbook.active

day_ranges = {
    "понедельник": (12, 19),
    "вторник": (23, 30),
    "среда": (34, 41),
    "четверг": (51, 52),
    "пятница": (56, 63),
    "суббота": (67, 74)
}

group_columns = {
    "кнт-1": [3, 5, 6],
    "кнт-2": [3, 8, 9],
    "кнт-3": [3, 11, 12],
    "кнт-4": [3, 17, 18],
    "кнт-5": [3, 20, 21],
    "кнт-6": [3, 23, 24],
    "кнт-7": [3, 29, 30],
    "кнт-8": [3, 32, 33],
    "кнт-9": [3, 35, 36]
}
def get_schedule_date(selected_day):
    """Возвращает дату для выбранного дня недели."""
    today = datetime.today()
    selected_day_num = list(day_ranges.keys()).index(selected_day) 
    days_difference = (selected_day_num - today.weekday()) % 7 
    
    if days_difference == 0: 
        return today.strftime("%d.%m.%Y")
    elif days_difference < 0: 
        days_difference += 7
    
    schedule_date = today + timedelta(days=days_difference)
    return schedule_date.strftime("%d.%m.%Y")

selected_group = None 
selected_group_message_id = None

@bot.message_handler(commands=['start'])
def main(message):
    current_message_id = message.message_id  
    markup = types.InlineKeyboardMarkup()
    markup.add(types.InlineKeyboardButton('24-кнт', callback_data='24-кнт'))
    sent_message = bot.reply_to(message, 'Привет, Выбери свой поток!', reply_markup=markup)

@bot.callback_query_handler(func=lambda callback: True)
def handle_callback(callback):
    global selected_group, selected_group_message_id
    message = callback.message
    chat_id = message.chat.id
    message_id = message.message_id
    if callback.data == '24-кнт':
        markup = types.InlineKeyboardMarkup(row_width=2) # Кнопки в два ряда
        back_button = types.InlineKeyboardButton('⏪ Назад', callback_data='back_to_start')
        markup.add(*[types.InlineKeyboardButton(group_name.upper(), callback_data=group_name) for group_name in group_columns], back_button)
        bot.edit_message_text('Выбери группу:', chat_id=chat_id, message_id=message_id, reply_markup=markup)
    elif callback.data == 'back_to_start': # обработка кнопки назад 
        markup = types.InlineKeyboardMarkup()
        markup.add(types.InlineKeyboardButton('24-кнт', callback_data='24-кнт'))
        bot.edit_message_text('Привет, Выбери свой поток!', chat_id=chat_id, message_id=message_id, reply_markup=markup)

    elif callback.data in group_columns:
        selected_group = callback.data
        msg = bot.send_message(chat_id, f"Вы выбрали группу {selected_group.upper()}") 
        selected_group_message_id = msg.message_id

        markup = types.InlineKeyboardMarkup(row_width=2) 
        back_button = types.InlineKeyboardButton('⏪ Назад', callback_data='24-кнт')
        markup.add(*[types.InlineKeyboardButton(day_name.capitalize(), callback_data=day_name) for day_name in day_ranges], back_button)
        bot.edit_message_text('Выбери день недели:', chat_id=chat_id, message_id=message_id, reply_markup=markup)
    
    elif callback.data in day_ranges and selected_group is not None:
        start_row, end_row = day_ranges[callback.data]
        selected_columns = group_columns[selected_group]
    
        rezult = ""
        for row_num in range(start_row, end_row + 1):
            vrem = ''
            empty_row = False
            for col_num in selected_columns:
                cell = sheet.cell(row=row_num, column=col_num)
                if cell.value is None:
                    empty_row = True
                    break

            if not empty_row:
                for col_num in selected_columns:
                    cell = sheet.cell(row=row_num, column=col_num)
                    cell_value = cell.value
                    vrem += str(cell_value) + " " 
                
            if not re.findall(r"\d{2}\.\d{2}", vrem) and not re.search(r"\d{1,2}:\d{2} - \d{1,2}:\d{2}\s*с\s*(\d{2}\.\d{2})", vrem):
                rezult += vrem + "\n\n"
            else:
                # Если даты есть, обрабатываем как и раньше:
                if re.search(r"\d{1,2}:\d{2} - \d{1,2}:\d{2}\s*с\s*(\d{2}\.\d{2})", vrem):
                    match = re.search(r"\d{1,2}:\d{2} - \d{1,2}:\d{2}\s*с\s*(\d{2}\.\d{2})", vrem)
                    date_str = match.group(1)
                    if is_date_past(date_str):  
                        rezult += vrem + "\n\n"
                elif re.findall(r"\d{2}\.\d{2}", vrem):
                    dates_in_row = re.findall(r"\d{2}\.\d{2}", vrem)
                    if dates_in_row and is_date_in_current_week(",".join(dates_in_row)):
                        rezult += vrem + "\n\n" 
        rezult = re.sub(r'\n{2,}', '\n\n', rezult)    
        day_markup = types.InlineKeyboardMarkup(row_width=2)        
        day_markup.add(*[types.InlineKeyboardButton(day_name.capitalize(), callback_data=day_name) for day_name in day_ranges])
        schedule_date_str = get_schedule_date(callback.data)
        if rezult.strip():
            bot.send_message(chat_id, f"📅 Дата: {schedule_date_str}\n{rezult}", reply_markup=day_markup)       
        else:
            bot.send_message(chat_id, f"📅 Дата: {schedule_date_str}\n\nНет данных для отображения", reply_markup=day_markup)
        bot.delete_message(chat_id, message_id)   
        if selected_group_message_id is not None:
            bot.delete_message(chat_id, selected_group_message_id)
            selected_group_message_id = None    

bot.polling(non_stop=True)

